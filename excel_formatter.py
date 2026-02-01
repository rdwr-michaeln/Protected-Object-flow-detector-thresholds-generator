"""
Excel report formatter for styling and highlighting threshold violations.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import CONFIG


class ExcelReportFormatter:
    """Handles Excel report formatting and styling."""
    
    def __init__(self, workbook_path: str):
        """Initialize the formatter with a workbook."""
        self.wb = load_workbook(workbook_path)
        self.ws = self.wb.active
    
    def format_report(self, dataframe: pd.DataFrame) -> None:
        """Apply all formatting to the Excel report."""
        self._insert_title_row()
        self._add_section_titles()
        self._format_section_titles()
        self._format_header_row()
        self._highlight_threshold_violations(dataframe)
        self._auto_adjust_column_widths()
        
    def _insert_title_row(self) -> None:
        """Insert a row at the top for section titles."""
        self.ws.insert_rows(1)
    
    def _add_section_titles(self) -> None:
        """Add section titles for Current Thresholds and Max Values."""
        # Merge cells for "Current Thresholds" (columns C to J)
        self.ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)
        self.ws.cell(row=1, column=3).value = "Current Thresholds"
        
        # Merge cells for "Max value for last 7 days" (columns K to R)
        self.ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=18)
        self.ws.cell(row=1, column=11).value = "Max value for last 7 days"
    
    def _format_section_titles(self) -> None:
        """Format the section title cells."""
        title_font = Font(bold=True, size=14)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format "Current Thresholds" title
        current_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        current_cell = self.ws.cell(row=1, column=3)
        current_cell.fill = current_fill
        current_cell.font = title_font
        current_cell.alignment = title_alignment
        
        # Format "Max value for last 7 days" title
        max_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        max_cell = self.ws.cell(row=1, column=11)
        max_cell.fill = max_fill
        max_cell.font = title_font
        max_cell.alignment = title_alignment
    
    def _format_header_row(self) -> None:
        """Format the column headers."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in self.ws[2]:  # Row 2 is now the header row
            cell.font = header_font
            cell.fill = header_fill
    
    def _highlight_threshold_violations(self, dataframe: pd.DataFrame) -> None:
        """Highlight rows where activation thresholds are below 80% of max values."""
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for row_idx in range(3, self.ws.max_row + 1):  # Start from row 3 (first data row)
            if self._should_highlight_row(row_idx):
                for col_idx in range(1, len(dataframe.columns) + 1):
                    self.ws.cell(row=row_idx, column=col_idx).fill = red_fill
    
    def _should_highlight_row(self, row_idx: int) -> bool:
        """Check if a row should be highlighted based on threshold violations."""
        threshold = CONFIG['THRESHOLD_PERCENTAGE']
        
        # Column mapping for comparison: (activation_col, max_col, description)
        comparisons = [
            (2, 10, "TCP Activation Mbps vs TCP Max BPS"),     # TCP Mbps
            (3, 11, "TCP Activation PPS vs TCP Max PPS"),      # TCP PPS
            (4, 12, "UDP Activation Mbps vs UDP Max BPS"),     # UDP Mbps
            (5, 13, "UDP Activation PPS vs UDP Max PPS"),      # UDP PPS
            (6, 14, "ICMP Activation Mbps vs ICMP Max BPS"),   # ICMP Mbps
            (7, 15, "ICMP Activation PPS vs ICMP Max PPS"),    # ICMP PPS
            (8, 16, "Total Activation Mbps vs Total Max BPS"), # Total Mbps
            (9, 17, "Total Activation PPS vs Total Max PPS"),  # Total PPS
        ]
        
        try:
            for activation_col, max_col, description in comparisons:
                activation_val = self._get_cell_value(row_idx, activation_col)
                max_val = self._get_float_value(row_idx, max_col)
                
                # Skip comparison if activation threshold is not configured
                if self._is_value_not_configured(activation_val):
                    continue
                
                activation_val_float = self._convert_to_float(activation_val)
                
                # Only highlight if activation is below threshold percentage of max
                if activation_val_float < max_val * threshold:
                    return True
            return False
        except Exception as e:
            print(f"Error in highlighting logic for row {row_idx}: {e}")
            return False
    
    def _is_value_not_configured(self, value) -> bool:
        """Check if a threshold value is not configured (empty, null, or zero)."""
        if value in [None, '', 0, '0', 'N/A', 'n/a']:
            return True
        try:
            # Consider very small values as not configured
            float_val = float(value)
            return float_val <= 0
        except (ValueError, TypeError):
            return True
    
    def _get_cell_value(self, row: int, col: int):
        """Get raw cell value without conversion."""
        return self.ws.cell(row=row, column=col).value
    
    def _convert_to_float(self, value) -> float:
        """Convert value to float, return 0 if conversion fails."""
        try:
            if value in [None, '']:
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _get_float_value(self, row: int, col: int) -> float:
        """Safely get a float value from a cell."""
        value = self.ws.cell(row=row, column=col).value
        if value in [None, '']:
            return 0.0
        return float(value)
    
    def _auto_adjust_column_widths(self) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx in range(1, self.ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = max_length + 2
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def save(self, filename: str) -> None:
        """Save the formatted workbook."""
        self.wb.save(filename)